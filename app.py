import streamlit as st
import datetime
import pytz
import pandas as pd
from db.database import get_events_by_date
from db.database import get_all_events
from core.reminder_engine import get_due_assignments
try:
    from core.email_reminder import send_email_reminders
    EMAIL_ENABLED = True
except:
    EMAIL_ENABLED = False
from tools.add_event_tool import add_event_tool
from tools.find_free_time_tool import find_free_time
from tools.add_assignment_tool import add_assignment_tool

from core.agent_controller import run_agent
from core.semester_manager import apply_semester_template

from core.calendar_service import (
    list_upcoming_events,
    delete_event,
    event_exists_on_date,
)

from core.assignment_manager import (
    get_assignments,
    mark_assignment_complete,
)

from db.database import initialize_database, delete_local_event


# ---------------- PAGE CONFIG ----------------

st.set_page_config(
    page_title="Academic Scheduling Agent",
    page_icon="📅",
    layout="wide",
)

initialize_database()

# ---------------- HEADER ----------------

st.title("📅 Academic Scheduling Agent")
st.markdown("AI-powered academic planner with smart scheduling & conflict management.")
st.divider()

# ---------------- REMINDERS ----------------

reminders = get_due_assignments()

if reminders:
    st.warning("⚠️ Upcoming Assignment Deadlines")

    today = datetime.date.today()

    for r in reminders:

        if isinstance(r, dict):
            aid = r["id"]
            title = r["title"]
            subject = r.get("subject", "General")
            due = r["due_date"]
            status = r.get("status", "")
        else:
            aid, title, subject, due = r
            status = ""

        subject = subject or "General"
        due_date = datetime.date.fromisoformat(due)

        if status == "past_due" or due_date < today:
            label = "⚠️ Past Due"
            show_delete = True
        elif status == "today" or due_date == today:
            label = "⏰ Due Today"
            show_delete = False
        else:
            label = "⏳ Due Tomorrow"
            show_delete = False

        col1, col2 = st.columns([5, 1])

        with col1:
            st.write(f"• {title} ({subject}) — {label}")

        with col2:
            if show_delete:
                if st.button("❌", key=f"rem_{aid}"):
                    mark_assignment_complete(aid)
                    st.rerun()

st.divider()

# ---------------- NAVIGATION ----------------

st.sidebar.header("Navigation")

pages = [
    "AI Assistant",
    "Dashboard",
    "Create Event",
    "Find Free Time",
    "Assignments",
]

page = st.sidebar.radio("Go to", pages)


# ==================================================
# 🤖 AI ASSISTANT (CLEAN MODE)
# ==================================================

if page == "AI Assistant":

    st.subheader("🤖 AI Scheduling Assistant")

    if st.button("📚 Apply Semester Template"):
        result = apply_semester_template(datetime.date.today().isoformat())
        st.success(result)

    st.divider()
    # 🔥 MEMORY STATE
    if "pending_action" not in st.session_state:
        st.session_state.pending_action = None
    query = st.text_input(
        "Enter your request",
        placeholder="e.g. schedule math exam tomorrow at 10am"
    )

    if st.button("Run AI"):

        if not query.strip():
            st.error("❌ Please enter a request")
        else:
            # 🔥 HANDLE FOLLOW-UP MEMORY
            if st.session_state.pending_action == "schedule":
                query = f"schedule {query}"
                st.session_state.pending_action = None

            response = run_agent(query)

            # 🔥 DETECT FOLLOW-UP QUESTIONS FROM AI
            if isinstance(response, str):
                if "what would you like to schedule" in response.lower():
                    st.session_state.pending_action = "schedule"

            st.divider()

            # 🔥 CLEAN OUTPUT BLOCK
            with st.container(border=True):
                st.markdown("### 🤖 Result")
                st.markdown(response)
# ==================================================
# 📊 DASHBOARD
# ==================================================

elif page == "Dashboard":

    st.subheader("📊 Schedule Dashboard")

    # ==================================================
    # 📧 REMINDERS
    # ==================================================
    st.subheader("📧 Reminders")

    if st.button("Send Reminder Email Now"):
        try:
            if EMAIL_ENABLED:
                send_email_reminders(
                    st.secrets["EMAIL_USER"],
                    st.secrets["EMAIL_PASS"],
                    st.secrets["USER_EMAIL"]
                )
            else:
                st.warning("📧 Email reminders not configured.")
        except Exception as e:
            st.error(f"❌ Email failed: {str(e)}")

    st.divider()

    IST = pytz.timezone("Asia/Kolkata")

    # ==================================================
    # 🌐 GOOGLE EVENTS (SAFE)
    # ==================================================
    try:
        events = list_upcoming_events()
    except Exception:
        events = []

    if events:

        st.subheader("🌐 Google Events")

        grouped = {}

        for event in events:

            start = event.get("start", {})

            if "dateTime" in start:
                dt = datetime.datetime.fromisoformat(
                    start["dateTime"].replace("Z", "+00:00")
                ).astimezone(IST)
                date_key = dt.date()
                time_str = dt.strftime("%I:%M %p")

            elif "date" in start:
                date_key = datetime.date.fromisoformat(start["date"])
                time_str = "All Day"
            else:
                continue

            grouped.setdefault(date_key, []).append((event, time_str))

        for date_key in sorted(grouped.keys()):

            st.markdown(f"### 📅 {date_key.strftime('%d %b %Y')}")

            for event, time_str in grouped[date_key]:

                with st.container(border=True):

                    col1, col2 = st.columns([4, 1])

                    with col1:
                        st.markdown(f"**{event.get('summary', 'Untitled')}**  \n🕒 {time_str}")

                    with col2:
                        if st.button("🗑", key=f"google_{event['id']}"):

                            success = delete_event(event["id"])

                            if success:
                                delete_local_event(event["id"])
                                st.success("Deleted")
                            else:
                                st.error("Delete failed (Google auth?)")

                            st.rerun()

    # ==================================================
    # 💾 LOCAL EVENTS (ALWAYS SHOW)
    # ==================================================
    local_events = get_all_events()

    if local_events:

        st.subheader("💾 Local Events")

        grouped = {}

        for title, date, start, end in local_events:
            # 🔥 FIX: normalize date before parsing
            year, month, day = date.split("-")
            date = f"{year}-{int(month):02d}-{int(day):02d}"

            date_obj = datetime.date.fromisoformat(date)
            grouped.setdefault(date_obj, []).append((title, start, end))

        for date_key in sorted(grouped.keys()):

            st.markdown(f"### 📅 {date_key.strftime('%d %b %Y')}")

            for idx, (title, start, end) in enumerate(grouped[date_key]):

                event_id = f"local_{title}_{date_key}_{start}"

                with st.container(border=True):

                    col1, col2 = st.columns([4, 1])

                    with col1:
                        st.markdown(f"**{title}**  \n🕒 {start} - {end}")

                    with col2:
                        if st.button("🗑", key=f"local_{idx}_{title}_{date_key}_{start}"):
                            delete_local_event(event_id)
                            st.success("Deleted")
                            st.rerun()

    # ==================================================
    # EMPTY STATE
    # ==================================================
    if not events and not local_events:
        st.info("No upcoming events.")

    st.divider()
    # ==================================================
    # 📤 EXPORT SCHEDULE (FINAL)
    # ==================================================

    st.subheader("📤 Export Schedule")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime

    export_data = []

    def format_dt(dt_str):
        try:
            dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
            return dt.strftime("%d-%m-%Y %H:%M")
        except:
            return dt_str

    # GOOGLE EVENTS
    for e in events or []:
        start = e.get("start", {})
        end = e.get("end", {})

        start_val = format_dt(start.get("dateTime") or start.get("date", ""))
        end_val = format_dt(end.get("dateTime") or end.get("date", ""))

        export_data.append([
            e.get("summary", "Untitled"),
            start_val,
            end_val,
            "Google"
        ])

    # LOCAL EVENTS
    for title, date, start, end in local_events:
        try:
            y, m, d = date.split("-")
            date = f"{y}-{int(m):02d}-{int(d):02d}"
        except:
            pass

        export_data.append([
            title,
            f"{date} {start}",
            f"{date} {end}",
            "Local"
        ])

    # CREATE EXCEL
    if export_data:
        wb = Workbook()
        ws = wb.active
        ws.title = "Schedule"

        ws.append(["Title", "Start", "End", "Type"])

        for row in export_data:
            ws.append(row)

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        st.download_button(
            "📥 Download Excel",
            buffer,
            file_name="schedule.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        st.info("No data available to export.")

# ==================================================
# CREATE EVENT
# ==================================================

elif page == "Create Event":

    st.subheader("➕ Create Event")

    title = st.text_input("Title")
    date = st.date_input("Date")
    start_time = st.time_input("Start Time")
    end_time = st.time_input("End Time")

    if st.button("Create"):

        if not title.strip():
            st.error("Title required")

        elif end_time <= start_time:
            st.error("Invalid time range")

        elif event_exists_on_date(title, date.isoformat()):
            st.error("Event already exists")

        else:
            result = add_event_tool(
                title,
                date.isoformat(),
                start_time.strftime("%H:%M"),
                end_time.strftime("%H:%M"),
            )

            if "⚠️" in result:
                st.warning(result)
                st.info("Try another time → Check Free Time below 👇")

                free = find_free_time(date=date.isoformat())
                st.markdown(free)

            elif "❌" in result:
                st.error(result)

            else:
                st.success(result)


# ==================================================
# FIND FREE TIME
# ==================================================

elif page == "Find Free Time":

    st.subheader("🕒 Free Time Finder")

    date = st.date_input("Select Date")

    if st.button("Check"):
        result = find_free_time(date=date.isoformat())
        st.markdown(result)


# ==================================================
# ASSIGNMENTS
# ==================================================

elif page == "Assignments":

    st.subheader("📚 Assignments")

    title = st.text_input("Title")
    subject = st.text_input("Subject")
    due_date = st.date_input("Due Date")
    priority = st.selectbox("Priority", ["auto", "low", "medium", "high"])

    if st.button("Add Assignment"):

        if not title.strip():
            st.error("Title required")

        else:
            result = add_assignment_tool(
                title=title,
                subject=subject,
                due_date=due_date.isoformat(),
                priority=priority,
            )
            st.success(result)

    st.divider()

    assignments = get_assignments()

    for a in assignments:

        with st.container(border=True):

            col1, col2 = st.columns([4, 1])

            with col1:
                st.markdown(
                    f"**{a['title']}** ({a['subject']})  \n🔥 {a['priority'].upper()}  \n⏰ {a['due_date']}"
                )

            with col2:
                if st.button("✔", key=f"a_{a['id']}"):
                    mark_assignment_complete(a['id'])
                    st.rerun()


# ---------------- FOOTER ----------------

st.divider()
st.caption("Production-ready Academic Scheduling Agent 🚀")